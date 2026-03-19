"""Collect individual MEP roll-call voting data from the European Parliament.

Fetches RCV XML files, maps each vote item to the vote_id from the main
Excel table, enriches with MEP details, and outputs in Supabase
meps_rcv_votes format.

Usage:
    python rcv_collector.py --date 2026-03-12           # Process specific date
    python rcv_collector.py --date 2026-03-12 --dry-run # Parse and compare only
    python rcv_collector.py --lookback 60               # Last 60 days
    python rcv_collector.py --build-mep-cache           # Build MEP cache from CSV
    python rcv_collector.py --export-csv                # Export DB to CSV
    python rcv_collector.py --stats                     # Show DB statistics
    python rcv_collector.py --verify 2026-03-12         # Compare against Supabase CSV
"""

from __future__ import annotations

import argparse
import csv
import json
import logging
import sqlite3
import time
from datetime import date, datetime, timedelta
from pathlib import Path

import requests

from ep_scraper.config import (
    CACHE_DIR, EXCEL_PATH, HEADERS, PROJECT_DIR,
    REQUEST_DELAY, REQUEST_TIMEOUT, TERM_START,
)
from ep_scraper.date_discovery import fetch_vot_xml
from ep_scraper.rcv_parser import MepVote, parse_rcv_xml
from ep_scraper.vot_parser import parse_vot_xml

log = logging.getLogger("rcv_collector")

# ── Paths & URLs ────────────────────────────────────────────────────────────
DB_PATH = PROJECT_DIR / "mep_votes.db"
RCV_CACHE_DIR = CACHE_DIR / "rcv"
MEP_CACHE_PATH = PROJECT_DIR / "mep_cache.json"

RCV_XML_URL = (
    "https://www.europarl.europa.eu/doceo/document/"
    "PV-10-{date}-RCV_EN.xml"
)

# ── Political group code → full name ───────────────────────────────────────
POLITICAL_GROUP_MAP = {
    "ECR": "European Conservatives and Reformists Group",
    "ESN": "Europe of Sovereign Nations Group",
    "NI": "Non-attached Members",
    "PfE": "Patriots for Europe Group",
    "PPE": "Group of the European People's Party (Christian Democrats)",
    "Renew": "Renew Europe Group",
    "S&D": "Group of the Progressive Alliance of Socialists and Democrats in the European Parliament",
    "The Left": "The Left group in the European Parliament - GUE/NGL",
    "Verts/ALE": "Group of the Greens/European Free Alliance",
}

# ── Position mapping ───────────────────────────────────────────────────────
POSITION_TO_VALUE = {"For": "+", "Against": "-", "Abstention": "0"}


# ── MEP cache ──────────────────────────────────────────────────────────────

def load_mep_cache() -> dict[str, dict]:
    """Load MEP details cache: {pers_id: {mep_name, country, political_group, national_party}}."""
    if MEP_CACHE_PATH.exists():
        with open(MEP_CACHE_PATH, encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_mep_cache(cache: dict[str, dict]) -> None:
    with open(MEP_CACHE_PATH, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)
    log.info("Saved MEP cache with %d entries to %s", len(cache), MEP_CACHE_PATH)


def build_mep_cache_from_csv(csv_path: Path) -> None:
    """Build MEP cache from a CSV export of Supabase meps_rcv_votes.

    The CSV should have columns: website_id, mep_name, country,
    political_group, national_party (plus others, ignored).
    Export from Supabase with:
        SELECT DISTINCT website_id, mep_name, country, political_group, national_party
        FROM ep_plenary_votes.meps_rcv_votes;
    """
    cache: dict[str, dict] = {}
    with open(csv_path, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            wid = str(row["website_id"])
            cache[wid] = {
                "mep_name": row["mep_name"],
                "country": row.get("country", ""),
                "political_group": row.get("political_group", ""),
                "national_party": row.get("national_party", ""),
            }
    save_mep_cache(cache)
    print(f"Built MEP cache: {len(cache)} MEPs from {csv_path}")


# ── Vote ID mapping (from Excel) ──────────────────────────────────────────

def get_vote_ids_for_date(date_str: str) -> list[int]:
    """Get ordered list of vote_ids for a date from the Excel file.

    Args:
        date_str: YYYY-MM-DD format.

    Returns:
        Vote IDs in ascending order.
    """
    from openpyxl import load_workbook

    # Convert to DD.MM.YYYY for Excel matching
    dt = datetime.strptime(date_str, "%Y-%m-%d")
    date_ddmmyyyy = dt.strftime("%d.%m.%Y")

    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb["RCVs"]

    vote_ids: list[int] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Column A = vote_id (idx 0), Column D = date (idx 3)
        vid = row[0]
        row_date = row[3]
        if row_date is None or vid is None:
            continue

        # Handle both string and datetime date values
        if hasattr(row_date, "strftime"):
            row_date_str = row_date.strftime("%d.%m.%Y")
        else:
            row_date_str = str(row_date).strip()

        if row_date_str == date_ddmmyyyy:
            vote_ids.append(int(vid))

    wb.close()
    return sorted(vote_ids)


def get_vote_tally_map_for_date(date_str: str) -> dict[int, tuple[int, int, int]]:
    """Get {vote_id: (yes, no, abs)} for a date from Excel."""
    from openpyxl import load_workbook

    dt = datetime.strptime(date_str, "%Y-%m-%d")
    date_ddmmyyyy = dt.strftime("%d.%m.%Y")

    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb["RCVs"]

    result: dict[int, tuple[int, int, int]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        vid = row[0]
        row_date = row[3]
        if row_date is None or vid is None:
            continue

        if hasattr(row_date, "strftime"):
            row_date_str = row_date.strftime("%d.%m.%Y")
        else:
            row_date_str = str(row_date).strip()

        if row_date_str == date_ddmmyyyy:
            # Columns W=22, X=23, Y=24 (0-indexed)
            yes = int(row[22]) if row[22] is not None else 0
            no = int(row[23]) if row[23] is not None else 0
            abs_ = int(row[24]) if row[24] is not None else 0
            result[int(vid)] = (yes, no, abs_)

    wb.close()
    return result


def build_voting_id_to_vote_id_map(
    date_str: str,
    session: requests.Session,
) -> dict[str, int]:
    """Build mapping from RCV Identifier (=votingId) → vote_id.

    Strategy:
    1. Parse VOT XML to get ROLL_CALL votingIds in document order.
    2. Get vote_ids for the date from Excel in ascending order.
    3. Try tally-based matching first (most robust).
    4. Fall back to position-based matching if tallies diverge.
    """
    d = datetime.strptime(date_str, "%Y-%m-%d").date()

    # Get votingIds from VOT XML
    vot_xml = fetch_vot_xml(d, session)
    if not vot_xml:
        log.error("No VOT XML for %s — cannot build mapping", date_str)
        return {}

    vot_records = parse_vot_xml(vot_xml, date_str)
    voting_ids = [rec.voting_id for rec in vot_records if rec.voting_id]

    # Get vote_ids from Excel
    vote_ids = get_vote_ids_for_date(date_str)

    if not voting_ids:
        log.error("No votingIds found in VOT XML for %s", date_str)
        return {}

    if len(voting_ids) != len(vote_ids):
        log.warning(
            "Count mismatch for %s: %d votingIds vs %d vote_ids",
            date_str, len(voting_ids), len(vote_ids),
        )
        # Try tally-based matching
        return _match_by_tally(vot_records, date_str)

    # Counts match → try tally matching, fall back to position
    tally_map = get_vote_tally_map_for_date(date_str)

    # Build tally → vote_id lookup
    tally_to_vid: dict[tuple[int, int, int], int] = {}
    for vid, tally in tally_map.items():
        tally_to_vid[tally] = vid  # last one wins if dupes

    mapping: dict[str, int] = {}
    matched_by_tally = 0

    for i, rec in enumerate(vot_records):
        if not rec.voting_id:
            continue
        tally = (rec.yes or 0, rec.no or 0, rec.abstentions or 0)
        if tally in tally_to_vid:
            mapping[rec.voting_id] = tally_to_vid[tally]
            matched_by_tally += 1
        elif i < len(vote_ids):
            mapping[rec.voting_id] = vote_ids[i]

    if matched_by_tally > 0:
        log.info("Matched %d/%d by tally, rest by position for %s",
                 matched_by_tally, len(voting_ids), date_str)
    else:
        # Pure position-based matching
        for vid_id, vote_id in zip(voting_ids, vote_ids):
            mapping[vid_id] = vote_id
        log.info("Position-matched %d items for %s", len(mapping), date_str)

    return mapping


def _match_by_tally(
    vot_records: list,
    date_str: str,
) -> dict[str, int]:
    """Match votingIds to vote_ids using tallies when counts differ."""
    tally_map = get_vote_tally_map_for_date(date_str)
    tally_to_vid: dict[tuple[int, int, int], int] = {}
    for vid, tally in tally_map.items():
        tally_to_vid[tally] = vid

    mapping: dict[str, int] = {}
    for rec in vot_records:
        if not rec.voting_id:
            continue
        tally = (rec.yes or 0, rec.no or 0, rec.abstentions or 0)
        if tally in tally_to_vid:
            mapping[rec.voting_id] = tally_to_vid[tally]

    log.info("Tally-matched %d/%d items for %s",
             len(mapping), len([r for r in vot_records if r.voting_id]), date_str)
    return mapping


# ── RCV XML fetching ────────────────────────────────────────────────────────

def _rcv_url(d: date) -> str:
    return RCV_XML_URL.format(date=d.strftime("%Y-%m-%d"))


def _rcv_cache_path(d: date) -> Path:
    return RCV_CACHE_DIR / f"PV-10-{d.strftime('%Y-%m-%d')}-RCV_EN.xml"


def fetch_rcv_xml(d: date, session: requests.Session) -> str | None:
    """Download RCV XML for a date. Cached to disk under cache/rcv/."""
    cache_file = _rcv_cache_path(d)
    if cache_file.exists():
        log.debug("Using cached RCV XML for %s", d)
        return cache_file.read_text(encoding="utf-8")

    url = _rcv_url(d)
    try:
        r = session.get(url, timeout=REQUEST_TIMEOUT)
        if r.status_code == 200 and r.text.strip().startswith("<?xml"):
            RCV_CACHE_DIR.mkdir(parents=True, exist_ok=True)
            cache_file.write_text(r.text, encoding="utf-8")
            log.info("Downloaded RCV XML for %s (%d KB)", d, len(r.text) // 1024)
            return r.text
        log.debug("No RCV XML for %s (HTTP %d)", d, r.status_code)
        return None
    except requests.RequestException as e:
        log.warning("Failed to fetch RCV XML for %s: %s", d, e)
        return None


def check_rcv_exists(d: date, session: requests.Session) -> bool:
    """HEAD-check whether an RCV XML exists for a date."""
    if _rcv_cache_path(d).exists():
        return True
    try:
        r = session.head(_rcv_url(d), timeout=REQUEST_TIMEOUT, allow_redirects=True)
        return r.status_code == 200
    except requests.RequestException:
        return False


# ── Database ────────────────────────────────────────────────────────────────

def init_db(db_path: Path = DB_PATH) -> sqlite3.Connection:
    """Create SQLite DB matching Supabase meps_rcv_votes schema."""
    conn = sqlite3.connect(str(db_path))
    conn.execute("PRAGMA journal_mode=WAL")

    conn.executescript("""
        CREATE TABLE IF NOT EXISTS mep_votes (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            website_id      INTEGER NOT NULL,
            mep_name        TEXT NOT NULL,
            country         TEXT,
            political_group TEXT,
            national_party  TEXT,
            vote_id         INTEGER NOT NULL,
            vote_value      TEXT NOT NULL CHECK(vote_value IN ('+','-','0')),
            source_dataset  TEXT DEFAULT 'Plenary votes'
        );

        CREATE TABLE IF NOT EXISTS processed_dates (
            date            TEXT PRIMARY KEY,
            n_items         INTEGER,
            n_mep_votes     INTEGER,
            processed_at    TEXT
        );

        CREATE INDEX IF NOT EXISTS idx_mv_vote_id ON mep_votes(vote_id);
        CREATE INDEX IF NOT EXISTS idx_mv_website_id ON mep_votes(website_id);
        CREATE INDEX IF NOT EXISTS idx_mv_group ON mep_votes(political_group);
    """)
    conn.commit()
    return conn


def get_processed_dates(conn: sqlite3.Connection) -> set[str]:
    rows = conn.execute("SELECT date FROM processed_dates").fetchall()
    return {r[0] for r in rows}


def insert_votes(
    conn: sqlite3.Connection,
    rows: list[dict],
    date_str: str,
    n_items: int,
) -> None:
    """Insert formatted MEP vote rows."""
    conn.executemany(
        """INSERT INTO mep_votes
           (website_id, mep_name, country, political_group,
            national_party, vote_id, vote_value, source_dataset)
           VALUES (:website_id, :mep_name, :country, :political_group,
                   :national_party, :vote_id, :vote_value, :source_dataset)""",
        rows,
    )
    conn.execute(
        """INSERT OR REPLACE INTO processed_dates
           (date, n_items, n_mep_votes, processed_at)
           VALUES (?, ?, ?, ?)""",
        (date_str, n_items, len(rows), datetime.now().isoformat()),
    )
    conn.commit()


# ── Date discovery ──────────────────────────────────────────────────────────

def discover_rcv_dates(
    processed: set[str],
    start: date | None = None,
    end: date | None = None,
    session: requests.Session | None = None,
) -> list[date]:
    """Find plenary dates with RCV XML not yet processed."""
    if start is None:
        start = datetime.strptime(TERM_START, "%Y-%m-%d").date()
    if end is None:
        end = date.today()
    if session is None:
        session = requests.Session()
        session.headers.update(HEADERS)

    new_dates: list[date] = []
    d = start
    while d <= end:
        if d.weekday() >= 5:
            d += timedelta(days=1)
            continue
        date_str = d.strftime("%Y-%m-%d")
        if date_str not in processed and check_rcv_exists(d, session):
            log.info("Found new RCV XML for %s", date_str)
            new_dates.append(d)
        time.sleep(REQUEST_DELAY)
        d += timedelta(days=1)

    return sorted(new_dates)


# ── Per-date processing ─────────────────────────────────────────────────────

def process_date(
    d: date,
    session: requests.Session,
    conn: sqlite3.Connection | None,
    mep_cache: dict[str, dict],
    dry_run: bool = False,
) -> tuple[int, int]:
    """Process one plenary date: fetch RCV XML, map to vote_ids, format rows.

    Returns (n_items, n_mep_votes).
    """
    date_str = d.strftime("%Y-%m-%d")
    log.info("Processing %s", date_str)

    # Step 1: Build votingId → vote_id mapping
    vid_map = build_voting_id_to_vote_id_map(date_str, session)
    if not vid_map:
        log.warning("No votingId mapping for %s — skipping", date_str)
        return 0, 0

    # Step 2: Parse RCV XML
    rcv_xml = fetch_rcv_xml(d, session)
    if not rcv_xml:
        log.warning("No RCV XML for %s", date_str)
        return 0, 0

    mep_votes, _ = parse_rcv_xml(rcv_xml, date_str)

    # Step 3: Format rows for Supabase schema
    rows: list[dict] = []
    unmapped = set()
    unknown_meps = set()

    for mv in mep_votes:
        if mv.rcv_id not in vid_map:
            unmapped.add(mv.rcv_id)
            continue

        vote_id = vid_map[mv.rcv_id]
        pers_id = mv.pers_id
        vote_value = POSITION_TO_VALUE[mv.position]

        # MEP lookup
        mep_info = mep_cache.get(pers_id)
        if mep_info:
            mep_name = mep_info["mep_name"]
            country = mep_info["country"]
            political_group = mep_info["political_group"]
            national_party = mep_info["national_party"]
        else:
            unknown_meps.add(pers_id)
            # Fallback: use RCV XML data
            mep_name = mv.mep_name
            country = ""
            political_group = POLITICAL_GROUP_MAP.get(mv.political_group, mv.political_group)
            national_party = ""

        rows.append({
            "website_id": int(pers_id) if pers_id else 0,
            "mep_name": mep_name,
            "country": country,
            "political_group": political_group,
            "national_party": national_party,
            "vote_id": vote_id,
            "vote_value": vote_value,
            "source_dataset": "Plenary votes",
        })

    n_items = len(vid_map)

    if unmapped:
        log.warning("  %d RCV items not mapped to vote_ids: %s",
                    len(unmapped), sorted(unmapped)[:5])
    if unknown_meps:
        log.warning("  %d MEPs not in cache (using RCV XML fallback)", len(unknown_meps))

    if dry_run:
        _display_summary(date_str, rows, n_items, len(unknown_meps))
    elif conn is not None:
        insert_votes(conn, rows, date_str, n_items)
        log.info("Stored %d MEP votes (%d items) for %s", len(rows), n_items, date_str)

    return n_items, len(rows)


def _display_summary(
    date_str: str, rows: list[dict], n_items: int, n_unknown: int
) -> None:
    """Print summary for dry-run mode."""
    groups: dict[str, dict[str, int]] = {}
    for r in rows:
        grp = r["political_group"][:20]
        groups.setdefault(grp, {"+": 0, "-": 0, "0": 0})
        groups[grp][r["vote_value"]] += 1

    print(f"\n{'='*70}")
    print(f"  Date:            {date_str}")
    print(f"  Roll-call items: {n_items}")
    print(f"  Total MEP votes: {len(rows)}")
    print(f"  Unknown MEPs:    {n_unknown}")
    print(f"  {'-'*66}")
    print(f"  {'Group':<22} {'For(+)':>8} {'Agn(-)':>8} {'Abs(0)':>8} {'Total':>8}")
    print(f"  {'-'*66}")
    for grp in sorted(groups):
        g = groups[grp]
        total = g["+"] + g["-"] + g["0"]
        print(f"  {grp:<22} {g['+']:>8} {g['-']:>8} {g['0']:>8} {total:>8}")
    print(f"{'='*70}")

    # Show vote_id mapping sample
    seen_vids = sorted({r["vote_id"] for r in rows})
    print(f"  Vote IDs: {seen_vids[:5]} ... {seen_vids[-3:]}" if len(seen_vids) > 8
          else f"  Vote IDs: {seen_vids}")


# ── CSV export ──────────────────────────────────────────────────────────────

def export_csv(conn: sqlite3.Connection, output_dir: Path | None = None) -> None:
    """Export DB to CSV matching Supabase schema."""
    if output_dir is None:
        output_dir = PROJECT_DIR / "rcv_export"
    output_dir.mkdir(exist_ok=True)

    path = output_dir / "meps_rcv_votes.csv"
    cursor = conn.execute(
        """SELECT website_id, mep_name, country, political_group,
                  national_party, vote_id, vote_value, source_dataset
           FROM mep_votes ORDER BY vote_id, political_group, mep_name""")
    cols = [desc[0] for desc in cursor.description]
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(cols)
        writer.writerows(cursor)

    total = conn.execute("SELECT COUNT(*) FROM mep_votes").fetchone()[0]
    print(f"Exported {total:,} rows to {path}")


# ── Stats ───────────────────────────────────────────────────────────────────

def print_db_stats(conn: sqlite3.Connection) -> None:
    total = conn.execute("SELECT COUNT(*) FROM mep_votes").fetchone()[0]
    n_dates = conn.execute("SELECT COUNT(*) FROM processed_dates").fetchone()[0]
    if n_dates == 0:
        print("Database is empty.")
        return

    date_range = conn.execute(
        "SELECT MIN(date), MAX(date) FROM processed_dates"
    ).fetchone()
    n_meps = conn.execute(
        "SELECT COUNT(DISTINCT website_id) FROM mep_votes"
    ).fetchone()[0]
    n_votes = conn.execute(
        "SELECT COUNT(DISTINCT vote_id) FROM mep_votes"
    ).fetchone()[0]

    print(f"\n  Database: {DB_PATH}")
    print(f"  Dates:    {n_dates} plenary days ({date_range[0]} to {date_range[1]})")
    print(f"  Votes:    {n_votes:,} distinct vote_ids")
    print(f"  Records:  {total:,} individual MEP votes")
    print(f"  MEPs:     {n_meps} distinct")


# ── CLI ─────────────────────────────────────────────────────────────────────

def setup_logging(verbose: bool = False) -> None:
    level = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(
        level=level,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
        datefmt="%H:%M:%S",
    )


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Collect individual MEP roll-call voting data (Supabase format)",
    )
    parser.add_argument("--date", type=str, help="Process specific date (YYYY-MM-DD)")
    parser.add_argument("--lookback", type=int, default=30, help="Days to look back (default: 30)")
    parser.add_argument("--dry-run", action="store_true", help="Parse and show stats only")
    parser.add_argument("--export-csv", action="store_true", help="Export DB to CSV")
    parser.add_argument("--stats", action="store_true", help="Show DB statistics")
    parser.add_argument("--build-mep-cache", type=str, metavar="CSV_PATH",
                        help="Build MEP cache from Supabase CSV export")
    parser.add_argument("--verbose", "-v", action="store_true", help="Debug logging")

    args = parser.parse_args()
    setup_logging(args.verbose)

    # Build MEP cache from CSV export
    if args.build_mep_cache:
        build_mep_cache_from_csv(Path(args.build_mep_cache))
        return

    conn = init_db()

    if args.stats:
        print_db_stats(conn)
        conn.close()
        return

    if args.export_csv:
        export_csv(conn)
        conn.close()
        return

    # Load MEP cache
    mep_cache = load_mep_cache()
    if not mep_cache:
        log.warning("MEP cache is empty. Run --build-mep-cache first for full MEP details.")
        log.warning("Continuing with RCV XML fallback data (no country/national_party).")

    session = requests.Session()
    session.headers.update(HEADERS)

    # Determine dates
    processed = get_processed_dates(conn)
    log.info("Already processed %d dates", len(processed))

    dates_to_process: list[date] = []

    if args.date:
        d = datetime.strptime(args.date, "%Y-%m-%d").date()
        if d.strftime("%Y-%m-%d") in processed and not args.dry_run:
            log.info("Date %s already processed. Use --dry-run to re-parse.", d)
            conn.close()
            return
        dates_to_process = [d]
    else:
        log.info("Checking last %d days for new RCV data...", args.lookback)
        start = date.today() - timedelta(days=args.lookback)
        dates_to_process = discover_rcv_dates(processed, start=start, session=session)

    if not dates_to_process:
        log.info("No new dates to process")
        print_db_stats(conn)
        conn.close()
        return

    log.info("Dates to process: %s",
             [d.strftime("%Y-%m-%d") for d in dates_to_process])

    # Process dates
    total_items = 0
    total_mep_votes = 0

    for d in dates_to_process:
        n_items, n_votes = process_date(
            d, session,
            conn=None if args.dry_run else conn,
            mep_cache=mep_cache,
            dry_run=args.dry_run,
        )
        total_items += n_items
        total_mep_votes += n_votes
        time.sleep(REQUEST_DELAY)

    print(f"\nProcessed {len(dates_to_process)} date(s): "
          f"{total_items} items, {total_mep_votes:,} MEP votes")

    if not args.dry_run:
        print_db_stats(conn)

    conn.close()


if __name__ == "__main__":
    main()
