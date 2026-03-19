"""Write vote records to the existing Dorian Excel database.

Preserves existing formatting, formulas, and data.
Creates a timestamped backup before any write operation.
"""

from __future__ import annotations

import logging
import shutil
from datetime import datetime
from pathlib import Path

import openpyxl

from ep_scraper.config import COL, EXCEL_PATH, SHEET_VOTES

log = logging.getLogger(__name__)


# ── Helpers ──────────────────────────────────────────────────────────────────

def _backup(excel_path: Path) -> Path:
    """Create a timestamped backup of the Excel file."""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = excel_path.parent / f"{excel_path.stem}_backup_{ts}{excel_path.suffix}"
    shutil.copy2(excel_path, backup_path)
    log.info("Created backup: %s", backup_path)
    return backup_path


def _ddmmyyyy_to_iso(val: str) -> str:
    """Convert ``DD.MM.YYYY`` → ``YYYY-MM-DD``."""
    try:
        parts = val.strip().split(".")
        if len(parts) == 3:
            return f"{parts[2]}-{parts[1]}-{parts[0]}"
    except Exception:
        pass
    return val


def _normalize_am_no(val) -> str:
    """Normalize am_no values for dedup key comparison.

    Converts floats like ``15.0`` → ``"15"``, ``0.0``/``None`` → ``"0"``.
    """
    if val is None:
        return "0"
    try:
        n = int(float(val))
        return str(n)
    except (ValueError, TypeError):
        return str(val).strip()


# ── Read helpers ─────────────────────────────────────────────────────────────

def get_existing_dates(excel_path: Path | None = None) -> set[str]:
    """Read all unique dates from the Excel database.

    Returns set of ``YYYY-MM-DD`` strings (internal format).
    Column D stores dates as ``DD.MM.YYYY`` strings.
    """
    if excel_path is None:
        excel_path = EXCEL_PATH
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    ws = wb[SHEET_VOTES]
    dates: set[str] = set()
    for row in ws.iter_rows(min_row=2, max_col=COL["date"], values_only=True):
        val = row[COL["date"] - 1]  # 0-based index
        if val is None:
            continue
        if isinstance(val, datetime):
            dates.add(val.strftime("%Y-%m-%d"))
        else:
            iso = _ddmmyyyy_to_iso(str(val))
            if len(iso) == 10:
                dates.add(iso)
    wb.close()
    return dates


def get_last_vote_id(excel_path: Path | None = None) -> int:
    """Get the maximum Vote ID from column A."""
    if excel_path is None:
        excel_path = EXCEL_PATH
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    ws = wb[SHEET_VOTES]
    last_id = 0
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        val = row[0]
        if val is not None:
            try:
                last_id = max(last_id, int(float(val)))
            except (ValueError, TypeError):
                pass
    wb.close()
    return last_id


def get_last_file_number(excel_path: Path | None = None) -> int:
    """Get the maximum File number from column B."""
    if excel_path is None:
        excel_path = EXCEL_PATH
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    ws = wb[SHEET_VOTES]
    last_file = 0
    for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        val = row[1]  # column B = index 1
        if val is not None:
            try:
                last_file = max(last_file, int(float(val)))
            except (ValueError, TypeError):
                pass
    wb.close()
    return last_file


def get_existing_keys(
    excel_path: Path | None = None,
) -> set[tuple[str, str, str, str]]:
    """Get set of ``(date, title, subject, am_no)`` 4-tuples for deduplication.

    Reads columns D (date), E (title), Q (subject), S (am_no).
    """
    if excel_path is None:
        excel_path = EXCEL_PATH
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    ws = wb[SHEET_VOTES]
    keys: set[tuple[str, str, str, str]] = set()

    # We need columns up to S (19)
    for row in ws.iter_rows(min_row=2, max_col=COL["am_no"], values_only=True):
        if row[0] is None:  # skip empty rows (col A = vote_id)
            continue

        # Date (col D, index 3)
        date_val = row[COL["date"] - 1]
        if isinstance(date_val, datetime):
            date_str = date_val.strftime("%d.%m.%Y")
        else:
            date_str = str(date_val).strip() if date_val else ""

        # Title (col E, index 4)
        title = str(row[COL["title"] - 1]).strip() if row[COL["title"] - 1] else ""

        # Subject (col Q, index 16)
        subject = str(row[COL["subject"] - 1]).strip() if row[COL["subject"] - 1] else ""

        # Am No. (col S, index 18)
        am_no = _normalize_am_no(row[COL["am_no"] - 1])

        keys.add((date_str, title, subject, am_no))
    wb.close()
    return keys


# ── Write ────────────────────────────────────────────────────────────────────

def _get_last_row(ws) -> int:
    """Find the last used row in a non-read-only worksheet."""
    row = ws.max_row
    while row > 1:
        if ws.cell(row=row, column=1).value is not None:
            return row
        row -= 1
    return 1  # header only


def write_records(
    records: list[dict],
    excel_path: Path | None = None,
    backup: bool = True,
) -> int:
    """Append vote records to the Excel database.

    Each *record* dict has keys matching :data:`COL` names.  The
    ``order_of_vote`` column is written as a formula referencing
    column K (Code).

    Returns the number of records written.
    """
    if excel_path is None:
        excel_path = EXCEL_PATH

    if not records:
        log.info("No records to write")
        return 0

    if backup:
        _backup(excel_path)

    wb = openpyxl.load_workbook(excel_path)
    ws = wb[SHEET_VOTES]

    last_row = _get_last_row(ws)
    next_row = last_row + 1

    written = 0
    for rec in records:
        r = next_row + written

        # Write each column value
        for col_name, col_idx in COL.items():
            # order_of_vote is a formula — handled separately below
            if col_name == "order_of_vote":
                continue
            val = rec.get(col_name)
            if val is not None:
                ws.cell(row=r, column=col_idx, value=val)

        # Write Order of vote formula in column C
        ws.cell(
            row=r,
            column=COL["order_of_vote"],
            value=f"=IF(K{r}=K{r - 1},C{r - 1}+1,1)",
        )

        written += 1

    wb.save(excel_path)
    wb.close()
    log.info("Wrote %d records starting at row %d", written, next_row)
    return written
